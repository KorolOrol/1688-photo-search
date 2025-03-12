import pytest
from openpyxl import load_workbook
from unittest.mock import patch, MagicMock
import os

from project.backend.excel_export import export_orders

@pytest.fixture
def mock_db():
    with patch('psycopg2.connect') as mock_connect:
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_connect.return_value = mock_conn
        mock_conn.cursor.return_value = mock_cursor
        yield mock_cursor

def test_export_orders_creates_excel_file(mock_db):
    mock_db.fetchall.return_value = [
        ('user1@example.com', 'Product 1', 2, 10.0, 20.0),
        ('user2@example.com', 'Product 2', 1, 15.0, 15.0)
    ]

    export_orders()

    assert os.path.exists('orders.xlsx')

    wb = load_workbook('orders.xlsx')
    ws = wb.active

    assert ws['A1'].value == 'Пользователь'
    assert ws['B1'].value == 'Товар'
    assert ws['C1'].value == 'Количество'
    assert ws['D1'].value == 'Цена'
    assert ws['E1'].value == 'Итого'

    assert ws['A2'].value == 'user1@example.com'
    assert ws['B2'].value == 'Product 1'
    assert ws['C2'].value == 2
    assert ws['D2'].value == 10.0
    assert ws['E2'].value == 20.0

    assert ws['A3'].value == 'user2@example.com'
    assert ws['B3'].value == 'Product 2'
    assert ws['C3'].value == 1
    assert ws['D3'].value == 15.0
    assert ws['E3'].value == 15.0

    os.remove('orders.xlsx')

def test_export_orders_no_orders(mock_db):
    mock_db.fetchall.return_value = []

    export_orders()

    assert os.path.exists('orders.xlsx')

    wb = load_workbook('orders.xlsx')
    ws = wb.active

    assert ws['A1'].value == 'Пользователь'
    assert ws['B1'].value == 'Товар'
    assert ws['C1'].value == 'Количество'
    assert ws['D1'].value == 'Цена'
    assert ws['E1'].value == 'Итого'

    assert ws.max_row == 1

    os.remove('orders.xlsx')